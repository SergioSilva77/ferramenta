"""Módulo Excel para rpaflow."""

from typing import Any, Optional


class Excel:
    """Classe para operações com planilhas Excel."""

    def __init__(self, filepath: str = ""):
        self.filepath = filepath
        self._workbook = None

    def open(self) -> bool:
        """Abre uma planilha Excel."""
        try:
            import openpyxl
            self._workbook = openpyxl.load_workbook(self.filepath)
            return True
        except FileNotFoundError:
            import openpyxl
            self._workbook = openpyxl.Workbook()
            return True
        except Exception as e:
            raise FileError(f"Erro ao abrir planilha: {e}")

    def read(self, sheet: str, range: Optional[str] = None) -> list:
        """Lê dados de uma planilha."""
        if not self._workbook:
            raise FileError("Planilha não aberta. Chame open() primeiro.")

        ws = self._workbook[sheet]

        if range:
            return [[cell.value for cell in row] for row in ws[range]]
        else:
            return [[cell.value for cell in row] for row in ws.iter_rows()]

    def write(self, sheet: str, range: str, values: list) -> bool:
        """Escreve dados em uma planilha."""
        if not self._workbook:
            raise FileError("Planilha não aberta. Chame open() primeiro.")

        ws = self._workbook[sheet]
        start_row = ws[range].min_row
        start_col = ws[range].min_col

        for i, row in enumerate(values):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=value)

        return True

    def save(self, filepath: Optional[str] = None) -> bool:
        """Salva a planilha."""
        if not self._workbook:
            raise FileError("Planilha não aberta. Chame open() primeiro.")

        self._workbook.save(filepath or self.filepath)
        return True

    def close(self) -> None:
        """Fecha a planilha."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
